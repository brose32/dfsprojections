
import sys


import openpyxl

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC



PATH = 'C:\Program Files (x86)\chromedriver.exe'
driver = webdriver.Chrome(PATH)
driver.maximize_window()
my_url = "https://dailyroto.com/nfl-snaps-draftkings-fanduel-fantasy-2016/"
driver.get(my_url)
element = WebDriverWait(driver, 10).until(
    EC.presence_of_all_elements_located((By.CLASS_NAME, "ng-scope"))
)
end_week = driver.find_element_by_xpath("//div[@class='left']/select[@name='endWeek']")
end_week.click()
week = driver.find_element_by_xpath("//div[@class='left']/select[@name='endWeek']/option[@value='17']")
week.click()
WebDriverWait(driver, 10).until(
  EC.presence_of_all_elements_located((By.CLASS_NAME, "ng-scope"))
)
positions = ['QB', 'RB', 'TE', 'WR']
pos_button = driver.find_element_by_id('pos')
pos_button.click()

player_dict = {
    'QB':{},
    'RB':{},
    'WR':{},
    'TE':{}
}
for pos in positions:
    pos_button = driver.find_element_by_id('pos')
    pos_button.click()
    position = driver.find_element_by_xpath(f"//option[@value='{pos}']")
    position.click()
    element = WebDriverWait(driver, 10).until(
        EC.presence_of_all_elements_located((By.XPATH, "//tr/td/b[@class='ng-binding']"))
    )
    #time.sleep(3)
    rows = driver.find_elements_by_class_name("ng-scope")
    print(len(rows))
    for i in range(len(rows)-2):
        #position, team, player, g, snaps/team, snap%, pergame, week14, week15, week16
        # week17, rush%, target%, touch%, fp, persnap
        name = driver.find_element_by_xpath(f"//table/tbody/tr[{i+1}]/td[3]").text
        player_dict[pos][name] = {
            'name': name,
            'snap_per': float(driver.find_element_by_xpath(f"//table/tbody/tr[{i+1}]/td[6]").text[:-1])/100,
            'rush_per': float(driver.find_element_by_xpath(f"//table/tbody/tr[{i+1}]/td[12]").text[:-1])/100,
            'target_per': float(driver.find_element_by_xpath(f"//table/tbody/tr[{i+1}]/td[13]").text[:-1])/100,
            'touch_per': float(driver.find_element_by_xpath(f"//table/tbody/tr[{i+1}]/td[14]").text[:-1])/100
        }

driver.quit()
print(player_dict['WR']['Tyreek Hill'])

my_file = "C:\\Users\\brose32\\Documents\\" + sys.argv[1]
wb = openpyxl.load_workbook(my_file)
qb_snaps_sheet = wb.create_sheet("QB SNAPS")
qb_snaps_sheet.append(("NAME", "SNAP%", "RUSH%"))
rb_snaps_sheet = wb.create_sheet("RB SNAPS")
rb_snaps_sheet.append(("NAME", "SNAP%", "RUSH%", "TARGET%", "TOUCH%"))
wr_snaps_sheet = wb.create_sheet("WR SNAPS")
wr_snaps_sheet.append(("NAME", "SNAP%", "RUSH%", "TARGET%", "TOUCH%"))
te_snaps_sheet = wb.create_sheet("TE SNAPS")
te_snaps_sheet.append(("NAME", "SNAP%", "RUSH%", "TARGET%", "TOUCH%"))
for player in player_dict['QB'].values():
    qb_snaps_sheet.append((player['name'], player['snap_per'], player['rush_per']))
for player in player_dict['RB'].values():
    rb_snaps_sheet.append((player['name'],player['snap_per'], player['rush_per'], player['target_per'], player['touch_per']))
for player in player_dict['WR'].values():
    wr_snaps_sheet.append((player['name'],player['snap_per'], player['rush_per'], player['target_per'], player['touch_per']))
for player in player_dict['TE'].values():
    te_snaps_sheet.append((player['name'],player['snap_per'], player['rush_per'], player['target_per'], player['touch_per']))
wb.save(my_file)

print('completed')