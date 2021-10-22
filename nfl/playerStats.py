import re
import sys
import time

import openpyxl

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

PATH = 'C:\Program Files (x86)\chromedriver.exe'
driver = webdriver.Chrome(PATH)
driver.maximize_window()
my_url = "https://www.fantasypros.com/nfl/stats/qb.php"
driver.get(my_url)
start = time.time()
elements = WebDriverWait(driver, 10).until(
    EC.presence_of_all_elements_located((By.TAG_NAME, "td"))
)
rows = driver.find_elements_by_xpath("//tbody/tr")

qbStats = {}
for i in range(len(rows)):
    attempts = float(driver.find_element_by_xpath(f"//tbody/tr[{i+1}]/td[4]").text)
    rushes = float(driver.find_element_by_xpath(f"//tbody/tr[{i+1}]/td[11]").text)
    name = driver.find_element_by_xpath(f"//tbody/tr[{i+1}]/td[2]").text
    clean_name = re.search('.+?(?= \()', name)[0]
    if attempts == 0:
        attempts = 1
    if rushes == 0:
        rushes = 1
    qbStats[clean_name] = {
        "yd/a" : float(driver.find_element_by_xpath(f"//tbody/tr[{i+1}]/td[7]").text),
        "paTD%" : float(driver.find_element_by_xpath(f"//tbody/tr[{i+1}]/td[8]").text) / attempts,
        "int%" : float(driver.find_element_by_xpath(f"//tbody/tr[{i+1}]/td[9]").text) /
                 attempts,
        "yd/r": float(driver.find_element_by_xpath(f"//tbody/tr[{i+1}]/td[12]").text.replace(",",""))/
                rushes,
        "ruTD%": float(driver.find_element_by_xpath(f"//tbody/tr[{i+1}]/td[13]").text) / rushes
    }
print("qbs grabbed", time.time() - start)

rb_button = driver.find_element_by_xpath("//li/a[@href='rb.php']")
rb_button.click()

elements = WebDriverWait(driver, 10).until(
    EC.presence_of_all_elements_located((By.TAG_NAME, "td"))
)
rows = driver.find_elements_by_xpath("//tbody/tr")
rbStats = {}
for i in range(len(rows)):
    recepts = float(driver.find_element_by_xpath(f"//tbody/tr[{i+1}]/td[9]").text)
    name = driver.find_element_by_xpath(f"//tbody/tr[{i+1}]/td[2]").text
    clean_name = re.search('.+?(?= \()', name)[0]
    attempts = float(driver.find_element_by_xpath(f"//tbody/tr[{i+1}]/td[3]").text)
    ruTD = float(driver.find_element_by_xpath(f"//tbody/tr[{i+1}]/td[8]").text)
    recTD = float(driver.find_element_by_xpath(f"//tbody/tr[{i+1}]/td[13]").text)
    targets = float(driver.find_element_by_xpath(f"//tbody/tr[{i+1}]/td[10]").text)
    if recepts != 0:
        rbStats[clean_name] = {
        "yd/c" : float(driver.find_element_by_xpath(f"//tbody/tr[{i+1}]/td[12]").text),
        "catch%" : recepts / targets,
        "recTD%" : recTD / targets
        }
    else:
        rbStats[clean_name] = {
            "yd/c": 0,
            "catch%":0,
            "recTD%":0
        }
    if attempts != 0:
        rbStats[clean_name]["ruTD%"] = ruTD / attempts
        rbStats[clean_name]["yd/r"] = float(driver.find_element_by_xpath(f"//tbody/tr[{i+1}]/td[5]").text)
    else:
        rbStats[clean_name]["ruTD%"] = 0
        rbStats[clean_name]["yd/r"] = 0

print("rbs grabbed", time.time() - start)
button = driver.find_element_by_xpath("//li/a[@href='wr.php']")
button.click()
elements = WebDriverWait(driver, 10).until(
    EC.presence_of_all_elements_located((By.TAG_NAME, "td"))
)
rows = driver.find_elements_by_xpath("//tbody/tr")
wrStats = {}
for i in range(len(rows)):
    recepts = float(driver.find_element_by_xpath(f"//tbody/tr[{i+1}]/td[3]").text)
    name = driver.find_element_by_xpath(f"//tbody/tr[{i+1}]/td[2]").text
    clean_name = re.search('.+?(?= \()', name)[0]
    attempts = float(driver.find_element_by_xpath(f"//tbody/tr[{i + 1}]/td[10]").text)
    ruTD = float(driver.find_element_by_xpath(f"//tbody/tr[{i + 1}]/td[12]").text)
    recTD = float(driver.find_element_by_xpath(f"//tbody/tr[{i + 1}]/td[9]").text)
    targets = float(driver.find_element_by_xpath(f"//tbody/tr[{i + 1}]/td[4]").text)
    if recepts != 0:
        wrStats[clean_name] = {
        "yd/c" : float(driver.find_element_by_xpath(f"//tbody/tr[{i+1}]/td[6]").text),
        "catch%" : recepts / targets,
        "recTD%" : recTD / targets
        }
    else:
        wrStats[clean_name] = {
            "yd/c": 0,
            "catch%":0,
            "recTD%":0
        }
    if attempts != 0:
        wrStats[clean_name]["ruTD%"] = ruTD / attempts
        wrStats[clean_name]["yd/r"] = float(driver.find_element_by_xpath(f"//tbody/tr[{i+1}]/td[11]").text) / attempts
    else:
        wrStats[clean_name]["ruTD%"] = 0
        wrStats[clean_name]["yd/r"] = 0

button = driver.find_element_by_xpath("//li/a[@href='te.php']")
button.click()
print("wrs grabbed", time.time() - start)
elements = WebDriverWait(driver, 10).until(
    EC.presence_of_all_elements_located((By.TAG_NAME, "td"))
)
rows = driver.find_elements_by_xpath("//tbody/tr")
teStats = {}
for i in range(len(rows)):
    recepts = float(driver.find_element_by_xpath(f"//tbody/tr[{i+1}]/td[3]").text)
    name = driver.find_element_by_xpath(f"//tbody/tr[{i+1}]/td[2]").text
    clean_name = re.search('.+?(?= \()', name)[0]
    attempts = float(driver.find_element_by_xpath(f"//tbody/tr[{i + 1}]/td[10]").text)
    tds = float(driver.find_element_by_xpath(f"//tbody/tr[{i + 1}]/td[12]").text) + float(driver.find_element_by_xpath(f"//tbody/tr[{i + 1}]/td[9]").text)
    targets = float(driver.find_element_by_xpath(f"//tbody/tr[{i + 1}]/td[4]").text)
    if recepts != 0:
        teStats[clean_name] = {
        "yd/c" : float(driver.find_element_by_xpath(f"//tbody/tr[{i+1}]/td[6]").text),
        "catch%" : recepts / targets,
        "TD%" : tds / (targets + attempts)
        }
    else:
        teStats[clean_name] = {
            "yd/c": 0,
            "catch%":0,
            "TD%":0
        }
print("tes grabbed", print(time.time() - start))
driver.quit()


my_file = "C:\\Users\\brose32\\Documents\\" + sys.argv[1]
wb = openpyxl.load_workbook(my_file)
qb_sheet = wb.create_sheet("QB STATS")
qb_sheet.append(("NAME", "Y/A", "PaTDper", "INT%", "Y/Ru", "RuTD%"))
rb_sheet = wb.create_sheet("RB STATS")
rb_sheet.append(("NAME", "Y/Ru", "RuTDper", "CATCH%", "Y/C", "RecTD%"))
wr_sheet = wb.create_sheet("WR STATS")
wr_sheet.append(("NAME", "Y/Ru", "RuTDper", "CATCH%", "Y/C", "RecTD%"))
te_sheet = wb.create_sheet("TE STATS")
te_sheet.append(("NAME", "CATCH%", "Y/C", "TDper"))
for qb in qbStats:
    qb_sheet.append((qb, qbStats[qb]["yd/a"], qbStats[qb]["paTD%"], qbStats[qb]["int%"], qbStats[qb]["yd/r"], qbStats[qb]["ruTD%"]))
for rb in rbStats:
    rb_sheet.append((rb, rbStats[rb]["yd/r"], rbStats[rb]["ruTD%"], rbStats[rb]["catch%"], rbStats[rb]["yd/c"], rbStats[rb]["recTD%"]))
for wr in wrStats:
    wr_sheet.append((wr, wrStats[wr]["yd/r"], wrStats[wr]["ruTD%"], wrStats[wr]["catch%"], wrStats[wr]["yd/c"], wrStats[wr]["recTD%"]))
for te in teStats:
    te_sheet.append((te, teStats[te]["catch%"], teStats[te]["yd/c"], teStats[te]["TD%"]))

wb.save(my_file)