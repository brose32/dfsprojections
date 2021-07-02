import sys

from bs4 import BeautifulSoup as soup
from urllib.request import urlopen
import openpyxl

#last 14 days stats
my_url = "https://www.fangraphs.com/leaders.aspx?pos=all&stats=sta&lg=all&qual=0&type=0&season=2021&month=2&season1=2021&ind=0&team=0&rost=0&age=0&filter=&players=0&startdate=2021-01-01&enddate=2021-12-31&page=1_300"

uClient = urlopen(my_url)
page_html = uClient.read()
uClient.close()
page_soup = soup(page_html, "html.parser")

table_body = page_soup.find("table", {"id" : "LeaderBoard1_dg1_ctl00"}).tbody
rows = table_body.findAll("tr")

pitcherLast14Dict = {}

for row in rows:
    #rank, name, team, w, l, era, g, gs, cg, sho, sv, hld, bs, ip, tbf, h, r, er, hr, bb, ibb, hbp, wp, bk, so
    rowData = row.findAll("td")
    pitcherLast14Dict[rowData[1].text.lstrip()] = {
        "name" : rowData[1].text.lstrip(),
        "team" : rowData[2].text,
        "era" : float(rowData[5].text),
        "g" : float(rowData[6].text),
        "gs" : float(rowData[7].text),
        "cg" : float(rowData[8].text),
        "sho" : float(rowData[9].text),
        "ip/g" : float(rowData[13].text) / float(rowData[6].text),
        "bf" : float(rowData[14].text),
        "h" : float(rowData[15].text),
        "h/g" : float(rowData[15].text) / float(rowData[6].text),
        "er/g" : float(rowData[17].text) / float(rowData[6].text),
        "bb" : float(rowData[19].text),
        "ibb" : float(rowData[20].text),
        "hbp" : float(rowData[21].text),
        "bb/g" : (float(rowData[19].text) + float(rowData[20].text) + float(rowData[21].text)) / float(rowData[6].text),
        "so/g" : float(rowData[24].text) / float(rowData[6].text)
    }
for key in pitcherLast14Dict:
    #base projection per inning
    pitcherLast14Dict[key]["last14FDproj"] = ((pitcherLast14Dict[key]["ip/g"] * 3) + (pitcherLast14Dict[key]["so/g"] * 3) \
                                             - (pitcherLast14Dict[key]["er/g"] * 3)) / pitcherLast14Dict[key]['ip/g']
    #project quality start
    if pitcherLast14Dict[key]["ip/g"] >= 6:
        pitcherLast14Dict[key]["last14FDproj"] += 2 / pitcherLast14Dict[key]["ip/g"]
    if pitcherLast14Dict[key]["er/g"] <= 3:
        pitcherLast14Dict[key]["last14FDproj"] += 2 / pitcherLast14Dict[key]["ip/g"]

my_url = "https://www.fangraphs.com/leaders.aspx?pos=all&stats=sta&lg=all&qual=10&type=0&season=2021&month=12&season1=2021&ind=0&team=0&rost=0&age=0&filter=&players=0&startdate=2021-01-01&enddate=2021-12-31&page=1_1500"
uClient = urlopen(my_url)
page_html = uClient.read()
uClient.close()
page_soup = soup(page_html, "html.parser")

table_body = page_soup.find("table", {"id" : "LeaderBoard1_dg1_ctl00"}).tbody
rows = table_body.findAll("tr")

pitcherLast3SeasonDict = {}

for row in rows:
    rowData = row.findAll("td")
    pitcherLast3SeasonDict[rowData[1].text.lstrip()] = {
        "name": rowData[1].text.lstrip(),
        "3season era": float(rowData[5].text),
        "3season g": float(rowData[6].text),
        "3season gs": float(rowData[7].text),
        "3season sho": float(rowData[8].text),
        "3season ip/g": float(rowData[13].text) / float(rowData[6].text),
        "3season bf": float(rowData[14].text),
        "3season h": float(rowData[15].text),
        "3season h/g": float(rowData[15].text) / float(rowData[6].text),
        "3season er/g": float(rowData[17].text) / float(rowData[6].text),
        "3season bb": float(rowData[19].text),
        "3season ibb": float(rowData[20].text),
        "3season hbp": float(rowData[21].text),
        "3season bb/g": (float(rowData[19].text) + float(rowData[20].text) + float(rowData[21].text)) / float(rowData[6].text),
        "3season so/g": float(rowData[24].text) / float(rowData[6].text)
    }

for key in pitcherLast3SeasonDict:
    #base projection per inning
    pitcherLast3SeasonDict[key]["last3SeasonFDproj"] = ((pitcherLast3SeasonDict[key]["3season ip/g"] * 3) + (pitcherLast3SeasonDict[key]["3season so/g"] * 3) \
                                             - (pitcherLast3SeasonDict[key]["3season er/g"] * 3)) / pitcherLast3SeasonDict[key]["3season ip/g"]
    #project quality start
    if pitcherLast3SeasonDict[key]["3season ip/g"] >= 6:
        pitcherLast3SeasonDict[key]["last3SeasonFDproj"] += 2 / pitcherLast3SeasonDict[key]["3season ip/g"]
    if pitcherLast3SeasonDict[key]["3season er/g"] <= 3:
        pitcherLast3SeasonDict[key]["last3SeasonFDproj"] += 2 / pitcherLast3SeasonDict[key]["3season ip/g"]

print(pitcherLast3SeasonDict["Carlos Rodon"])
my_file = "C:\\Users\\brose32\\Documents\\" + sys.argv[1]
wb = openpyxl.load_workbook(my_file)
pitcher_stats_sheet_last_14 = wb.create_sheet("PITCHER LAST 14")
pitcher_stats_sheet_last_14.append(("NAME", "LAST 14 FD AVG", "IP/G"))
for key in pitcherLast14Dict:
    pitcher_stats_sheet_last_14.append((key, pitcherLast14Dict[key]["last14FDproj"], pitcherLast14Dict[key]["ip/g"]))
pitcher_stats_sheet_last_3_season = wb.create_sheet("PITCHER L3 SEASON")
pitcher_stats_sheet_last_3_season.append(("NAME", "PITCHER L3 SEASON"))
for key in pitcherLast3SeasonDict:
    pitcher_stats_sheet_last_3_season.append((key, pitcherLast3SeasonDict[key]["last3SeasonFDproj"]))
wb.save(my_file)